#!/usr/bin/python

from lxml import html

import datetime
import json
import pprint
import re
import requests
import subprocess
import sys

try:
    from openpyxl.workbook import Workbook
    from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side, Style
    from openpyxl.cell import Cell
except ImportError:
    print('You need to execute:\n\tsudo pip install openpyxl')
    sys.exit(-1)

codeReviewSearchURL = """https://codereview.chromium.org/search?closed=1&owner=%s&reviewer=&cc=&repo_guid=&base=&project=&private=1&commit=1&created_before=&created_after=&modified_before=&modified_after=&order=&format=html&keys_only=False&with_messages=False&cursor=&limit=30"""
codeReviewURL = 'https://codereview.chromium.org/user/'
repositories = [
    {
        'name': 'blink',
        'path': '/home/vivekg/workspace/report-generation/blink',
        'type': 'git',
        'host': 'https://chromium.googlesource.com/',
        'guid': '82804f94bf4b4d551c864ff2ef44317633f76246'
    },
    {
        'name': 'chromium',
        'path': '/home/vivekg/workspace/report-generation/chromium',
        'type': 'git',
        'host': 'https://chromium.googlesource.com/',
        'guid': 'c14d891d44f0afff64e56ed7c9702df1d807b1ee'
    },
    {
        'name': 'skia',
        'path': '/home/vivekg/workspace/report-generation/skia',
        'type': 'git',
        'host': 'https://chromium.googlesource.com/',
        'guid': '586101c79b0490b50623e76c71a5fd67d8d92b08'
    },
    {
        'name': 'trace-viewer',
        'path': '/home/vivekg/workspace/report-generation/trace-viewer',
        'type': 'git',
        'host': 'https://github.com',
        'url': 'https://github.com/google/trace-viewer'
    },
    {
        'name': 'v8',
        'path': '/home/vivekg/workspace/report-generation/v8',
        'type': 'git',
        'host': 'https://chromium.googlesource.com/',
        'guid': '33f2fb0e53d135f0ee17cfccd9d993eb2a6f47de'
    }
]

authors = [
    { 'name': 'Abhijeet Kandalkar', 'email': 'abhijeet.k at samsung dot com' },
    { 'name': 'Ajay Berwal', 'email': 'ajay.berwal at samsung dot com' },
    { 'name': 'Akhil Teeka Dhananjaya', 'email': 'akhil.td at samsung dot com' },
    { 'name': 'Behara Mani Shyam Patro', 'email': 'behara.ms at samsung dot com' },
    { 'name': 'Gandhi Kishor Addanki', 'email': 'kishor.ag at samsung dot com' },
    { 'name': 'Ganesh Kamat', 'email': 'ganesh.kamat at samsung dot com' },
    { 'name': 'Kaja Mohaideen', 'email': 'kaja.m at samsung dot com' },
    { 'name': 'Karthik Gopalan', 'email': 'karthikg.g at samsung dot com' },
    { 'name': 'Kulajit Das', 'email': 'das.kulajit at samsung dot com' },
    { 'name': 'Mallikarjuna Narala', 'email': 'mallik.n at samsung dot com' },
    { 'name': 'Munukutla Subrahmanya Praveen', 'email': 'sataya.m at samsung dot com' },
    { 'name': 'Nikhil Sahni', 'email': 'nikhil.sahni at samsung dot com' },
    { 'name': 'Pavan Kumar Emani', 'email': 'pavan.e at samsung dot com' },
    { 'name': 'Prabhavathi Perumal', 'email': 'prabha.p at samsung dot com' },
    { 'name': 'Prashant Nevase', 'email': 'prashant.n at samsung dot com' },
    { 'name': 'Putturaju R', 'email': 'puttaraju.r at samsung dot com' },
    { 'name': 'Ravi Kasibhatla', 'email': 'r.kasibhatla at samsung dot com', 'username': 'kphanee' },
    { 'name': 'Shanmuga Pandi', 'email': 'shanmuga.m at samsung dot com' },
    { 'name': 'Siva Gunturi', 'email': 'siva.gunturi at samsung dot com' },
    { 'name': 'Sohan Jyoti Ghosh', 'email': 'sohan.jyoti at samsung dot com' },
    { 'name': 'Suchit Agarwal', 'email': 'a.suchit at samsung dot com' },
    { 'name': 'Sujith S S', 'email': 'sujiths.s at samsung dot com' },
    { 'name': 'Suyash Sengar', 'email': 'suyash.s at samsung dot com' },
    { 'name': 'Tanvir Rizvi', 'email': 'tanvir.rizvi at samsung dot com' },
    { 'name': 'Thanikassalam Kankayan', 'email': 'thanik.k at samsung dot com' },
    { 'name': 'Vivek Agrawal', 'email': 'vivek.s14 at samsung dot com' },
    { 'name': 'Vivek Galatage', 'email': ['vivek.vg at samsung dot com', ''] }
]

day = datetime.date.today() - datetime.timedelta(days=7)
start = day - datetime.timedelta(days = day.weekday())
end = start + datetime.timedelta(days = 6)
weekStart = start.strftime('%Y-%m-%d')
weekEnd = end.strftime('%Y-%m-%d')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'))

greyTheme = Style(
    border=thin_border,
    fill=PatternFill(fill_type='solid', start_color='FFD8D8D8'))

whiteTheme = Style(
    border=thin_border,
    fill=PatternFill(fill_type='solid', start_color='FFFFFFFF'))

wb = Workbook()

def beautifyWorksheet(sheet):
    sheet.column_dimensions['A'].width = 30
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[sheet.get_highest_row()].height = 30

    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        sheet.column_dimensions[col].width = 12

    header = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1']
    lastRow = str(sheet.get_highest_row())
    footer = ['A' + lastRow, 'B' + lastRow, 'C' + lastRow, 'D' + lastRow, 'E' + lastRow, 'F' + lastRow, 'G' + lastRow]

    for col in header:
        sheet[col].style = Style(
            alignment=Alignment(vertical='bottom', horizontal='center'),
            border=thin_border,
            fill=PatternFill(fill_type='solid', start_color='FFFFFF99'),
            font=Font(bold=True))

    for col in footer:
        sheet[col].style = Style(
            border=thin_border,
            fill=PatternFill(fill_type='solid', start_color='FFFFFF99'),
            font=Font(bold=True))

    for col in range(sheet.get_highest_column()):
        for row in range(1, sheet.get_highest_row() - 1):
            cell = sheet[str(chr(col + 65) + str(row + 1))]
            if (row % 2) == 0:
                cell.style = greyTheme
            else:
                cell.style = whiteTheme

def generateYearlyReport(authorData, year):
    thisYear = wb.create_sheet(index=1, title= year + ' Contributions')
    thisYear['A1'] = 'Name'
    thisYear['B1'] = 'Chromium'
    thisYear['C1'] = 'Blink'
    thisYear['D1'] = 'Trace-Viewer'
    thisYear['E1'] = 'Skia'
    thisYear['F1'] = 'V8'
    thisYear['G1'] = 'Total'
    for i in range(len(authorData)):
        author = authorData[i]
        index = str(i + 2)
        thisYear['A' + index] = author['name']

        if type(author['email']) == list:
            thisYear['A' + index].hyperlink = codeReviewSearchURL % author['email'][0]
        else:
            thisYear['A' + index].hyperlink = codeReviewSearchURL % author['email']

        if 'contributions' not in author:
            continue;

        contributions = author['contributions']

        repoToColumnTuples = [
            ('B', 'chromium'),
            ('C', 'blink'),
            ('D', 'trace-viewer'),
            ('E', 'skia'),
            ('F', 'v8'),
        ]

        for (col, repo) in repoToColumnTuples:
            if repo in contributions:
                if year not in contributions[repo]:
                    continue
                thisYear[col + index] = contributions[repo][year]

        thisYear['G' + index] = '=sum(b%s:f%s)' % (index, index)
    finalRowIndex = str(len(authorData) + 2)
    dataRange = '(%s2:%s' + str(len(authorData) + 1) + ')'
    thisYear['A' + finalRowIndex] = 'Total'

    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        thisYear[col + finalRowIndex] = '=sum' + (dataRange % (col, col))
    beautifyWorksheet(thisYear)

def generateReport(authorData):
    totals = wb.create_sheet(index=0, title='Total')
    totals['A1'] = 'Name'
    totals['B1'] = 'Chromium'
    totals['C1'] = 'Blink'
    totals['D1'] = 'Trace-Viewer'
    totals['E1'] = 'Skia'
    totals['F1'] = 'V8'
    totals['G1'] = 'Total'

    for i in range(len(authorData)):
        author = authorData[i]
        index = str(i + 2)
        totals['A' + index] = author['name']

        if type(author['email']) == list:
            totals['A' + index].hyperlink = codeReviewSearchURL % author['email'][0]
        else:
            totals['A' + index].hyperlink = codeReviewSearchURL % author['email']

        if 'contributions' not in author:
            continue;

        contributions = author['contributions']

        repoToColumnTuples = [
            ('B', 'chromium'),
            ('C', 'blink'),
            ('D', 'trace-viewer'),
            ('E', 'skia'),
            ('F', 'v8'),
        ]

        for (col, repo) in repoToColumnTuples:
            if repo in contributions:
                totals[col + index] = contributions[repo]['total']

        totals['G' + index] = '=sum(b%s:f%s)' % (index, index)
    finalRowIndex = str(len(authorData) + 2)
    dataRange = '(%s2:%s' + str(len(authorData) + 1) + ')'
    totals['A' + finalRowIndex] = 'Total'

    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        totals[col + finalRowIndex] = '=sum' + (dataRange % (col, col))
    beautifyWorksheet(totals)
    generateYearlyReport(authorData, str(datetime.date.today().year))
    wb.save('weeklyReport.xlsx')

def execute(cwd, command, verbose=True, progress=False):
    process = subprocess.Popen(command, shell=True, cwd=cwd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    processOutput = []
    while True:
        nextline = process.stdout.readline()
        if nextline == '' and process.poll() != None:
            break
        processOutput.append(nextline)
        if (verbose):
            sys.stdout.write(nextline)
            sys.stdout.flush()
        if (progress):
            sys.stdout.write('.')
            sys.stdout.flush()
    if (progress):
        print('')
    output = process.communicate()[0]
    exitCode = process.returncode

    if (exitCode == 0):
        return processOutput
    else:
        raise ProcessException(command, exitCode, output)

def getFormattedAuthorEmails():
    authorEmails = []
    for author in authors:
        if (type(author['email']) == list):
            authorEmails = authorEmails +  author['email']
        elif (type(author['email']) == str):
            authorEmails.append(author['email'])
    authorsList = []
    for email in authorEmails:
        authorsList.append('\(' + email + '\)')
    return '\|'.join(authorsList)

def parseGitLogOutput(repo, output, key):
    allContributions = {}
    for line in output:
        details = line.split()
        commits = int(details[0])
        retriedEmail = details[len(details) - 1]
        email = re.compile('((^<)(\w+\.?\w+@\w+.?\w+)(.*$))').match(retriedEmail).groups()[2]
        if email not in allContributions:
            allContributions[email] = commits
        else:
            allContributions[email] = allContributions[email] + commits
    if not len(allContributions):
        return

    for i in range(len(authors)):
        emailType = type(authors[i]['email'])
        commits = 0
        if emailType == list:
            for eachEmail in authors[i]['email']:
                if eachEmail in allContributions:
                    commits = commits + allContributions[eachEmail]
        elif emailType == str:
            email = authors[i]['email']
            if email in allContributions:
                commits = allContributions[email]
        if commits:
            repoName = repo['name']
            if 'contributions' not in authors[i]:
                authors[i]['contributions'] = {}
            if repoName not in authors[i]['contributions']:
                authors[i]['contributions'][repoName] = {}
            authors[i]['contributions'][repoName][key] = commits

formattedAuthors = getFormattedAuthorEmails()

TOTAL = 0
YEARLY = 1
WEEKLY = 2

global totalOnce
global yearOnce
global weeklyOnce


totalOnce = False
yearOnce = False
weeklyOnce = False

def populateContribution(duration, authors, year=None, begin=None, end=None):
    global totalOnce
    global yearOnce
    global weeklyOnce

    QUERY = [
        ('total', 'git shortlog -es --author="%s"'),
        ('%s', 'git shortlog -es --author="%s" --after="%s-01-01T00:00:00+00:00" --before="%s-12-31T23:59:59+00:00"'),
        ('weekly', 'git shortlog -es --author="%s" --after="%sT00:00:00+00:00" --before="%s-12-31T23:59:59+00:00"')]

    for repo in repositories:
        command = ''
        key = ''
        if duration == TOTAL:
            if not totalOnce:
                sys.stdout.write('\nCalculating total contributions so far, a big task you see!')
            totalOnce = True
            command = QUERY[TOTAL][1] % (authors)
            key = QUERY[TOTAL][0]
        elif duration == YEARLY:
            assert(year != None)
            if not yearOnce:
                sys.stdout.write('\nFiguring out what kept you busy all these years')
            yearOnce = True
            command = QUERY[YEARLY][1] % (authors, year, year)
            key = QUERY[YEARLY][0] % year
        elif duration == WEEKLY:
            assert(begin != None and end != None)
            if not weeklyOnce:
                sys.stdout.write('\nOpen source rocks!!! Lets check out what are you upto currently')
            weeklyOnce = True
            command = QUERY[WEEKLY][1] % (authors, begin, end)
            key = QUERY[WEEKLY][0]
        sys.stdout.write('.')
        sys.stdout.flush()
        output = execute(repo['path'], command, False)
        parseGitLogOutput(repo, output, key)

def updateRepositories():
    for repo in repositories:
        sys.stdout.write('Updating repository: ' + repo['name'])
        if (repo['type'] == 'git'):
            output = execute(repo['path'], 'git pull', False, True)
    print('')

def processRietveld(author, guid, begin, end):
    reitveldURL = 'https://codereview.chromium.org/search?closed=%s&owner=%s&repo_guid=%s&modified_after=%s&modified_before=%s&limit=30'
    assert(author != None and guid != None and begin != None and end != None)
    email = ''
    if type(author['email']) == list:
        email = author['email'][0]
    else:
        email = author['email']
    closedIssuesURL = reitveldURL % ('2', email, guid, weekStart, weekEnd)
    page = requests.get(closedIssuesURL)
    tree = html.fromstring(page.text)
    issueDetails = tree.xpath('//*[@class="subject"]/a/text()')
    closedIssues = []
    for i in xrange(0, len(issueDetails), 2):
        issueURL = 'https://codereview.chromium.org/' + issueDetails[i]
        issueTitle = issueDetails[i + 1].strip()
        closedIssues.append((issueURL, issueTitle))

    openIssuesURL = reitveldURL % ('3', email, guid, weekStart, weekEnd)
    page = requests.get(openIssuesURL)
    tree = html.fromstring(page.text)
    issueDetails = tree.xpath('//*[@class="subject"]/a/text()')
    openIssues = []
    for i in xrange(0, len(issueDetails), 2):
        issueURL = 'https://codereview.chromium.org/' + issueDetails[i]
        issueTitle = issueDetails[i + 1].strip()
        openIssues.append((issueURL, issueTitle))

    return (closedIssues, openIssues)

def processGithub(author, url, begin, end):
    assert(author['username'] != None)
    githubURL = url + '/pulls?q=is:pr+author:%s+is:%s+updated:%s..%s'
    #begin = '2014-11-01'
    #end = '2014-12-01'
    closedIssuesURL = githubURL % (author['username'], 'closed', begin, end)
    page = requests.get(closedIssuesURL)
    tree = html.fromstring(page.text)
    issueDetails = tree.xpath('//*[@class="issue-title-link js-navigation-open"]')
    closedIssues = []
    for issue in issueDetails:
        issueURL = 'https://github.com' + issue.get('href')
        issueTitle = issue.text.strip()
        closedIssues.append((issueURL, issueTitle))

    openIssuesURL = githubURL % (author['username'], 'open', begin, end)
    page = requests.get(openIssuesURL)
    tree = html.fromstring(page.text)
    openIssues = []
    issueDetails = tree.xpath('//*[@class="issue-title-link js-navigation-open"]')
    for issue in issueDetails:
        issueURL = 'https://github.com' + issue.get('href')
        issueTitle = issue.text.strip()
        openIssues.append((issueURL, issueTitle))

    return (closedIssues, openIssues)


def populateContributionDetails():
    sys.stdout.write('\nPreparing the contribution details, hang on!')
    for repo in repositories:
        repoName = repo['name']
        for i in xrange(len(authors)):
            author = authors[i]
            sys.stdout.write('.')
            sys.stdout.flush()
            if 'contributions' not in authors[i]:
                continue

            if repoName not in authors[i]['contributions']:
                continue

            assert('host' in repo)
            if repo['host'] == 'https://chromium.googlesource.com/':
                assert(repo['guid'] != None)
                issues = processRietveld(author, repo['guid'], weekStart, weekEnd)
                closedIssues = issues[0]
                openIssues = issues[1]
            elif repo['host'] == 'https://github.com':
                assert(repo['url'] != None)
                if 'username' in author:
                    issues = processGithub(author, repo['url'], weekStart, weekEnd)
                    closedIssues = issues[0]
                    openIssues = issues[1]

            if 'closed' not in authors[i]['contributions'][repoName]:
                authors[i]['contributions'][repoName]['closed'] = []
                if 'open' not in authors[i]['contributions'][repoName]:
                    authors[i]['contributions'][repoName]['open'] = []

                for issue in closedIssues:
                    newIssue = {};
                    newIssue[issue[0]] = issue[1]
                    authors[i]['contributions'][repoName]['closed'].append(newIssue);

                for issue in openIssues:
                    newIssue = {};
                    newIssue[issue[0]] = issue[1]
                    authors[i]['contributions'][repoName]['open'].append(newIssue);
    print('')

def main():
    # print 'Generating report for the week: %s - %s' % (weekStart, weekEnd)
    # updateRepositories()
    # populateContribution(TOTAL, formattedAuthors)
    # for year in range(2012, datetime.date.today().year + 1):
    #     populateContribution(YEARLY, formattedAuthors, year=str(year))
    # populateContribution(WEEKLY, formattedAuthors, begin=weekStart, end=weekEnd)
    # populateContributionDetails()

    # with open('weeklyReport.json', 'w') as jsonFile:
    #     json.dump(authors, jsonFile, sort_keys=True, indent=4)

    with open('weeklyReport.json') as jsonFile:
        authorData = json.loads(jsonFile.read())

    generateReport(authorData)

    #print('Hush!!! You guys are just awesome, kept me busy till now! See you next week!')

if __name__ == '__main__':
    sys.exit(main())
